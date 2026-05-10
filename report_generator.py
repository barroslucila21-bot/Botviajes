from database import get_flights_today, get_packages_today
from datetime import datetime
import logging, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)
OUTPUT_FILE = "/tmp/informe_buzios.xlsx"
PASAJERAS = 5
ARS_RATE = 1250

def fill(c): return PatternFill("solid", fgColor=c)
def brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def hf(color="FFFFFF", sz=10): return Font(name="Calibri", bold=True, color=color, size=sz)
def cf(sz=9, bold=False, color="1A1A2E"): return Font(name="Calibri", size=sz, bold=bold, color=color)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

AEROLINEAS = {
    "LATAM": {"puntualidad":"87%","cancelaciones":"Baja","equipaje":"23 kg incluido","estrellas":"LLLLL","color":"EBF5FB"},
    "Copa Airlines": {"puntualidad":"85%","cancelaciones":"Baja","equipaje":"23 kg incluido","estrellas":"LLLLO","color":"D6EAF8"},
    "Azul": {"puntualidad":"81%","cancelaciones":"Baja-Media","equipaje":"23 kg incluido","estrellas":"LLLLO","color":"D5F5E3"},
    "Gol": {"puntualidad":"78%","cancelaciones":"Media","equipaje":"Costo extra ~$25 USD","estrellas":"LLLOO","color":"FFF3CD"},
    "Aerolineas Arg.": {"puntualidad":"72%","cancelaciones":"Media-Alta","equipaje":"23 kg incluido","estrellas":"LLLOO","color":"FDEBD0"},
}

def dur_fmt(m):
    if not m: return "N/D"
    h, mm = divmod(int(m), 60)
    return f"{h}h {mm:02d}m"

def generate_daily_report(filename=OUTPUT_FILE):
    vuelos = get_flights_today()
    paquetes = get_packages_today()
    if not vuelos and not paquetes:
        log.warning("Sin datos en DB.")
        return None, None

    wb = Workbook()
    wb.remove(wb.active)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    # SOLAPA 1 - RESUMEN
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L1")
    ws["A1"] = f"TRAVEL ADVISOR - BUZIOS / RIO DE JANEIRO | {now}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill("0D1B2A"); ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 36

    if vuelos:
        pp = [v.get("price_usd",0) for v in vuelos if v.get("price_usd")]
        if pp:
            minp=min(pp); maxp=max(pp); avgp=round(sum(pp)/len(pp))
            mejor=min(vuelos, key=lambda v: v.get("price_usd",9999))
            directos=sum(1 for v in vuelos if v.get("stops",1)==0)
            ws.merge_cells("A2:L2")
            ws["A2"] = f"MIN: USD {minp:.0f}/pp (x5: USD {minp*5:.0f})  |  PROM: USD {avgp}/pp  |  MAX: USD {maxp:.0f}/pp  |  {len(vuelos)} vuelos  |  {directos} directos"
            ws["A2"].font = Font(name="Calibri", size=10, color="1E8449", bold=True)
            ws["A2"].fill = fill("D5F5E3"); ws["A2"].alignment = ctr(); ws.row_dimensions[2].height=22
            ws.merge_cells("A3:L3")
            aero=mejor.get("airline",""); fi=mejor.get("departure_date","")
            esc="Directo" if mejor.get("stops",1)==0 else "Escala"
            ws["A3"] = f"MEJOR: {aero} | {fi} | {esc} | USD {mejor.get('price_usd',0):.0f}/pp | Total x5: USD {mejor.get('price_usd',0)*5:.0f} | {mejor.get('baggage','N/D')}"
            ws["A3"].font = Font(name="Calibri", size=10, color="1B4F72", bold=True)
            ws["A3"].fill = fill("EBF5FB"); ws["A3"].alignment = lft(); ws.row_dimensions[3].height=22

            ws.merge_cells("A5:L5")
            ws["A5"] = "TOP 5 MAS BARATOS"
            ws["A5"].font = hf("FFFFFF",11); ws["A5"].fill = fill("2E86C1"); ws["A5"].alignment=ctr(); ws.row_dimensions[5].height=24

            hdrs=["#","Aerolinea","Origen","Destino","Fecha","Escala","Duracion","USD/pp","Total x5","Equipaje","Fuente","Ver"]
            for ci,h in enumerate(hdrs,1):
                c=ws.cell(6,ci,h); c.font=hf("FFFFFF",9); c.fill=fill("1B4F72"); c.alignment=ctr(); c.border=brd()
            ws.row_dimensions[6].height=24
            top5=sorted(vuelos,key=lambda v:v.get("price_usd",9999))[:5]
            for ri,v in enumerate(top5,7):
                bg="D5F5E3" if ri==7 else ("F2F3F4" if ri%2==0 else "FFFFFF")
                info=AEROLINEAS.get(v.get("airline",""),{})
                vals=[ri-6,v.get("airline",""),v.get("origin",""),v.get("destination",""),
                      v.get("departure_date",""),"Directo" if v.get("stops",1)==0 else "Escala",
                      dur_fmt(v.get("duration_min",0)),v.get("price_usd",0),v.get("price_usd",0)*PASAJERAS,
                      v.get("baggage","N/D"),v.get("fuente","N/D"),"Ver oferta"]
                for ci,val in enumerate(vals,1):
                    cell=ws.cell(ri,ci,val); cell.fill=fill(bg); cell.border=brd(); cell.alignment=ctr()
                    if ci in[8,9]:
                        cell.number_format='#,##0 "USD"'
                        cell.font=Font(name="Calibri",size=9,bold=(ri==7),color="1E8449" if ri==7 else "1A1A2E")
                    else: cell.font=cf(9,bold=(ri==7 and ci==2))
                url=v.get("url","https://www.google.com")
                ws.cell(ri,12).hyperlink=url
                ws.cell(ri,12).font=Font(name="Calibri",size=9,color="1155CC",underline="single")
                ws.row_dimensions[ri].height=15

    widths=[3,14,7,7,11,10,10,12,13,20,13,20]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A7"

    # SOLAPA 2 - VUELOS
    wv=wb.create_sheet("Vuelos"); wv.sheet_view.showGridLines=False
    wv.merge_cells("A1:P1")
    wv["A1"]=f"VUELOS - Buzios/Rio | {now}"
    wv["A1"].font=hf("FFFFFF",13); wv["A1"].fill=fill("0D1B2A"); wv["A1"].alignment=ctr(); wv.row_dimensions[1].height=34
    hdrs2=[("#",3),("Aerolinea",14),("Rating",8),("Orig",7),("Dest",7),("Fecha Ida",11),("Fecha Vuelta",12),
           ("Noches",7),("Escala",14),("Duracion",10),("USD/pp",12),("Total x5",13),("ARS est.",13),
           ("Equipaje",20),("Puntual",11),("Fuente",13),("Ver",22)]
    for ci,(lbl,w) in enumerate(hdrs2,1):
        c=wv.cell(2,ci,lbl); c.font=hf("FFFFFF",9); c.fill=fill("1B4F72"); c.alignment=ctr(); c.border=brd()
        wv.column_dimensions[get_column_letter(ci)].width=w
    wv.row_dimensions[2].height=28
    if vuelos:
        vord=sorted(vuelos,key=lambda v:v.get("price_usd",9999))
        minp2=min(v.get("price_usd",9999) for v in vord)
        for ri,v in enumerate(vord,3):
            p=v.get("price_usd",0) or 0; es_min=(p==minp2)
            bg="D5F5E3" if es_min else ("F2F3F4" if ri%2==0 else "FFFFFF")
            info=AEROLINEAS.get(v.get("airline",""),{})
            vals=[ri-2,v.get("airline",""),info.get("estrellas",""),v.get("origin",""),v.get("destination",""),
                  v.get("departure_date",""),v.get("return_date",""),6,
                  "Directo" if v.get("stops",1)==0 else "Escala",dur_fmt(v.get("duration_min",0)),
                  p,p*PASAJERAS,p*ARS_RATE,v.get("baggage","N/D"),info.get("puntualidad","N/D"),v.get("fuente","N/D"),None]
            for ci,val in enumerate(vals,1):
                cell=wv.cell(ri,ci,val); cell.fill=fill(bg); cell.border=brd()
                cell.alignment=lft() if ci in[2,15] else ctr()
                if ci in[11,12]: cell.number_format='#,##0 "USD"'; cell.font=Font(name="Calibri",size=9,bold=es_min,color="1E8449" if es_min else "1A1A2E")
                elif ci==13: cell.number_format='#,##0 "ARS"'; cell.font=cf(8,color="566573")
                else: cell.font=cf(9,bold=es_min and ci==2)
            url=v.get("url","https://www.google.com")
            lc=wv.cell(ri,17,"Ver oferta"); lc.hyperlink=url
            lc.font=Font(name="Calibri",size=9,color="1155CC",underline="single")
            lc.fill=fill(bg); lc.border=brd(); lc.alignment=ctr()
            wv.row_dimensions[ri].height=15
    wv.freeze_panes="A3"

    # SOLAPA 3 - PAQUETES
    wp=wb.create_sheet("Paquetes"); wp.sheet_view.showGridLines=False
    wp.merge_cells("A1:J1")
    wp["A1"]=f"PAQUETES Vuelo+Hotel | {now}"
    wp["A1"].font=hf("FFFFFF",13); wp["A1"].fill=fill("1E8449"); wp["A1"].alignment=ctr(); wp.row_dimensions[1].height=34
    phdrs=[("#",3),("Hotel",24),("Fecha Ida",11),("Fecha Vuelta",12),("Aerolinea",13),("USD paq.",13),("USD/pp",13),("Equipaje",20),("Fuente",13),("Ver",22)]
    for ci,(lbl,w) in enumerate(phdrs,1):
        c=wp.cell(2,ci,lbl); c.font=hf("FFFFFF",9); c.fill=fill("27AE60"); c.alignment=ctr(); c.border=brd()
        wp.column_dimensions[get_column_letter(ci)].width=w
    wp.row_dimensions[2].height=28
    if paquetes:
        pord=sorted(paquetes,key=lambda p:p.get("price_usd",9999))
        minpq=min(p.get("price_usd",9999) for p in pord)
        for ri,p in enumerate(pord,3):
            pu=p.get("price_usd",0) or 0; es_min=(pu==minpq)
            bg="D5F5E3" if es_min else ("F2F3F4" if ri%2==0 else "FFFFFF")
            info=AEROLINEAS.get(p.get("airline",""),{})
            vals=[ri-2,p.get("hotel_name","N/D"),p.get("fecha_ida",""),p.get("fecha_vuelta",""),
                  p.get("airline","N/D"),pu,round(pu/PASAJERAS,0) if pu else 0,
                  info.get("equipaje","N/D"),p.get("fuente","N/D"),None]
            for ci,val in enumerate(vals,1):
                cell=wp.cell(ri,ci,val); cell.fill=fill(bg); cell.border=brd()
                cell.alignment=lft() if ci==2 else ctr()
                if ci in[6,7]: cell.number_format='#,##0 "USD"'; cell.font=Font(name="Calibri",size=9,bold=es_min and ci in[6,7],color="1E8449" if es_min else "1A1A2E")
                else: cell.font=cf(9)
            url=p.get("url","https://www.despegar.com")
            lc=wp.cell(ri,10,"Ver"); lc.hyperlink=url
            lc.font=Font(name="Calibri",size=9,color="1155CC",underline="single")
            lc.fill=fill(bg); lc.border=brd(); lc.alignment=ctr()
            wp.row_dimensions[ri].height=15
    wp.freeze_panes="A3"

    # SOLAPA 4 - HISTORICO
    wh=wb.create_sheet("Historico"); wh.sheet_view.showGridLines=False
    wh.merge_cells("A1:H1")
    wh["A1"]="HISTORICO + ASESORIA | EZE/AEP -> GIG/SDU"
    wh["A1"].font=hf("FFFFFF",13); wh["A1"].fill=fill("E67E22"); wh["A1"].alignment=ctr(); wh.row_dimensions[1].height=34
    hhdrs=[("Aerolinea",16),("Rating",8),("Puntual",11),("Cancelac.",13),("Equipaje",20),("Min hist.",13),("Max hist.",13),("Precio hoy",14),("Recomendacion",30)]
    for ci,(lbl,w) in enumerate(hhdrs,1):
        c=wh.cell(2,ci,lbl); c.font=hf("FFFFFF",9); c.fill=fill("E67E22"); c.alignment=ctr(); c.border=brd()
        wh.column_dimensions[get_column_letter(ci)].width=w
    wh.row_dimensions[2].height=28
    notas={"LATAM":"La mas recomendada. Mejor balance precio-calidad. Equipaje 23kg incluido.",
           "Copa Airlines":"Muy confiable. Escala en Panama. Buen servicio a bordo.",
           "Azul":"Buena low-cost. Tiene vuelos directos desde Argentina.",
           "Gol":"La mas barata pero equipaje se paga APARTE (~USD 25). Solo con carry-on.",
           "Aerolineas Arg.":"Sale de Aeroparque. Peor historial de puntualidad del grupo."}
    hist_ref={"LATAM":(200,610),"Copa Airlines":(180,550),"Azul":(170,510),"Gol":(155,500),"Aerolineas Arg.":(210,630)}
    for ri,(aero_n,info) in enumerate(AEROLINEAS.items(),3):
        vv=[v for v in vuelos if v.get("airline","")==aero_n] if vuelos else []
        min_hoy=min((v.get("price_usd",9999) for v in vv),default=None)
        pmin,pmax=hist_ref.get(aero_n,(200,600))
        bg=info.get("color","FFFFFF")
        hoy_str=f"USD {min_hoy:.0f}" if min_hoy else "Sin datos"
        vals=[aero_n,info["estrellas"],info["puntualidad"],info["cancelaciones"],info["equipaje"],pmin,pmax,hoy_str,notas.get(aero_n,"")]
        for ci,val in enumerate(vals,1):
            cell=wh.cell(ri,ci,val); cell.fill=fill(bg); cell.border=brd()
            cell.alignment=lft() if ci in[1,5,9] else ctr()
            if ci in[6,7]: cell.number_format='#,##0 "USD"'; cell.font=cf(9)
            else: cell.font=cf(9)
        wh.row_dimensions[ri].height=20
    wh.freeze_panes="A3"

    # SOLAPA 5 - POR FECHA
    wf=wb.create_sheet("Por Fecha"); wf.sheet_view.showGridLines=False
    wf.merge_cells("A1:G1")
    wf["A1"]="ANALISIS POR SEMANA - Cuando conviene viajar?"
    wf["A1"].font=hf("FFFFFF",13); wf["A1"].fill=fill("7D3C98"); wf["A1"].alignment=ctr(); wf.row_dimensions[1].height=34
    fhdrs=[("Semana",13),("Vuelta",12),("Vuelos",8),("Min USD/pp",13),("Prom USD/pp",13),("Max USD/pp",13),("Veredicto",18)]
    for ci,(lbl,w) in enumerate(fhdrs,1):
        c=wf.cell(2,ci,lbl); c.font=hf("FFFFFF",9); c.fill=fill("7D3C98"); c.alignment=ctr(); c.border=brd()
        wf.column_dimensions[get_column_letter(ci)].width=w
    wf.row_dimensions[2].height=28
    semanas=[("2026-01-03","2026-01-09"),("2026-01-10","2026-01-16"),("2026-01-17","2026-01-23"),("2026-01-24","2026-01-30")]
    best2=9999
    if vuelos:
        for fi,fv in semanas:
            vv=[v for v in vuelos if v.get("departure_date","")==fi]
            if vv: best2=min(best2,min(v.get("price_usd",9999) for v in vv))
    for ri,(fi,fv) in enumerate(semanas,3):
        vv=[v for v in vuelos if v.get("departure_date","")==fi] if vuelos else []
        pp=[v.get("price_usd",0) for v in vv if v.get("price_usd")]
        fi_f=f"{fi[8:]}/{fi[5:7]}/{fi[:4]}"; fv_f=f"{fv[8:]}/{fv[5:7]}/{fv[:4]}"
        if pp: mn=min(pp); mx=max(pp); av=round(sum(pp)/len(pp))
        else: mn=mx=av=0
        if mn==best2 and pp: vtxt="LA MAS BARATA"
        elif mn<300 and pp: vtxt="Muy barata"
        elif mn<450 and pp: vtxt="Normal"
        elif pp: vtxt="Cara"
        else: vtxt="Sin datos"
        bg="E8DAEF" if (pp and mn==best2) else ("F2F3F4" if ri%2==0 else "FFFFFF")
        vals=[fi_f,fv_f,len(vv),mn or "-",av or "-",mx or "-",vtxt]
        for ci,val in enumerate(vals,1):
            cell=wf.cell(ri,ci,val); cell.fill=fill(bg); cell.border=brd(); cell.alignment=ctr()
            if ci in[4,5,6] and isinstance(val,(int,float)) and val>0:
                cell.number_format='#,##0 "USD"'
                cell.font=Font(name="Calibri",size=10,bold=(ci==4 and pp and mn==best2),color="1E8449" if (ci==4 and pp and mn==best2) else "1A1A2E")
            elif ci==7:
                cm={"LA MAS BARATA":"1E8449","Muy barata":"1E8449","Normal":"E67E22","Cara":"C0392B"}
                cell.font=Font(name="Calibri",size=10,bold=True,color=cm.get(val,"566573"))
            else: cell.font=cf(10,bold=(pp and mn==best2))
        wf.row_dimensions[ri].height=20
    wf.freeze_panes="A3"

    wb.save(filename)
    log.info(f"Excel guardado: {filename}")
    return filename, _mensaje(vuelos, paquetes)

def _mensaje(vuelos, paquetes):
    now=datetime.now().strftime("%d/%m/%Y %H:%M")
    if not vuelos: return f"Reporte {now} - Sin datos. Usa /buscar."
    pp=[v.get("price_usd",0) for v in vuelos if v.get("price_usd")]
    if not pp: return f"Reporte {now} - Sin precios aun."
    minp=min(pp); maxp=max(pp); avgp=round(sum(pp)/len(pp))
    mejor=min(vuelos,key=lambda v:v.get("price_usd",9999))
    directos=sum(1 for v in vuelos if v.get("stops",1)==0)
    top3=sorted(vuelos,key=lambda v:v.get("price_usd",9999))[:3]
    top3_txt=""
    for i,v in enumerate(top3,1):
        esc="Directo" if v.get("stops",1)==0 else "Escala"
        top3_txt+=f"  {i}. {v.get('airline','')} | {v.get('departure_date','')} | {esc} | <b>USD {v.get('price_usd',0):.0f}/pp</b> (x5: USD {v.get('price_usd',0)*5:.0f})\n"
    semanas={"2026-01-03":"3-9 ene","2026-01-10":"10-16 ene","2026-01-17":"17-23 ene","2026-01-24":"24-30 ene"}
    mejor_sem=""; mejor_p_sem=9999
    for fi,lbl in semanas.items():
        vv=[v for v in vuelos if v.get("departure_date","")==fi]
        if vv:
            mp=min(v.get("price_usd",9999) for v in vv)
            if mp<mejor_p_sem: mejor_p_sem=mp; mejor_sem=lbl
    if minp<200: veredicto="PRECIO HISTORICO BAJO - COMPRA YA!"
    elif minp<280: veredicto="Precio muy bueno - conviene comprar"
    elif minp<380: veredicto="Precio normal para la ruta"
    else: veredicto="Precio alto - espera una baja"
    esc_m="Directo" if mejor.get("stops",1)==0 else "Escala"
    return (f"<b>REPORTE DIARIO DE VUELOS</b>\n"
            f"{now} | Buzios/Rio | 5 pasajeras | 6 noches\n"
            f"------------------------\n\n"
            f"<b>PRECIOS HOY</b>\n"
            f"  Minimo: <b>USD {minp:.0f}/pp</b> -> x5: <b>USD {minp*5:.0f}</b>\n"
            f"  Promedio: USD {avgp}/pp | Maximo: USD {maxp:.0f}/pp\n\n"
            f"<b>MEJOR OPCION</b>\n"
            f"  {mejor.get('airline','')} | {mejor.get('departure_date','')} | {esc_m}\n"
            f"  <b>USD {mejor.get('price_usd',0):.0f}/pp - x5: USD {mejor.get('price_usd',0)*5:.0f}</b>\n"
            f"  {mejor.get('baggage','N/D')} | Via {mejor.get('fuente','N/D')}\n\n"
            f"<b>TOP 3 MAS BARATOS</b>\n{top3_txt}\n"
            f"<b>STATS</b>: {len(vuelos)} vuelos | {directos} directos | {len(paquetes)} paquetes\n"
            f"Semana mas barata: <b>{mejor_sem}</b> (desde USD {mejor_p_sem:.0f}/pp)\n\n"
            f"<b>VEREDICTO</b>: {veredicto}\n\n"
            f"Excel adjunto con detalles, historico y links para comprar.")
